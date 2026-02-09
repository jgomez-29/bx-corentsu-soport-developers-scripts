"""
Servicio para leer y procesar archivos Excel de boletas.

Flujo:
    1. Leer Excel de entrada → extraer HESCode y demás datos
    2. Hacer match con respuesta de API por HESCode
    3. Agregar columnas: BOLETA y DETALLE_ERRORES
    4. Escribir Excel de salida

Formato del Excel:
    - Debe tener una columna con el header que contenga "HES" (ej: "HESCode", "HES Code", etc.)
    - Se agregan 2 columnas al final:
        * BOLETA: BTECode si exitoso, 0 si error
        * DETALLE_ERRORES: errorDetails.message o status si error, vacío si exitoso
"""

import openpyxl
from openpyxl import Workbook
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import re


def translate_error_message(error_msg: str) -> str:
    """
    Traduce y simplifica mensajes de error comunes de la API a español.

    Args:
        error_msg: Mensaje de error en inglés

    Returns:
        Mensaje traducido y simplificado en español
    """
    if not error_msg:
        return ""

    # Patrón: communeName: Commune 'X' does not exist in region 'Y'
    match = re.search(r"communeName: Commune '([^']+)' does not exist in region '([^']+)'", error_msg)
    if match:
        commune = match.group(1)
        region = match.group(2)
        return f"Comuna '{commune}' no existe en la región '{region}'"

    # Patrón: sii_error: Failed to extract hidden fields: ...
    if "sii_error" in error_msg.lower() and "failed to extract hidden fields" in error_msg.lower():
        return "Error del SII: No se pudieron extraer campos ocultos del formulario"

    # Patrón: PROVIDER_IDENTIFIER_FORMAT_NOT_VALID
    if error_msg == "PROVIDER_IDENTIFIER_FORMAT_NOT_VALID":
        return "Formato de RUT del proveedor inválido"

    # Otros estados conocidos
    error_translations = {
        "BTE_CREATE_ERROR": "Error al crear BTE",
        "NO_ENCONTRADO_EN_API": "No encontrado en respuesta de API",
        "ERROR_DESCONOCIDO": "Error desconocido",
    }

    return error_translations.get(error_msg, error_msg)


def find_hes_column(worksheet) -> Optional[int]:
    """
    Busca la columna que contiene "HES" en el header.

    Args:
        worksheet: Worksheet de openpyxl

    Returns:
        Índice de la columna (1-based) o None si no se encuentra
    """
    for cell in worksheet[1]:  # Primera fila (headers)
        if cell.value and "HES" in str(cell.value).upper():
            return cell.column
    return None


def read_excel_data(excel_path: str) -> Tuple[Workbook, List[Dict[str, Any]], int]:
    """
    Lee el Excel y extrae los datos.

    Args:
        excel_path: Ruta al archivo Excel

    Returns:
        Tupla: (Workbook, lista de registros con {row_index, hes_code, ...}, columna HES)

    Raises:
        FileNotFoundError: Si no existe el archivo
        ValueError: Si no se encuentra la columna HES
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel: {excel_path}")

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    hes_column = find_hes_column(ws)
    if not hes_column:
        raise ValueError(f"No se encontró una columna con 'HES' en el header del Excel: {excel_path}")

    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
        hes_value = row[hes_column - 1].value  # Convert to 0-based index
        if hes_value is not None:
            try:
                hes_code = int(hes_value)
                records.append({
                    "row_index": row_idx,
                    "hes_code": hes_code,
                })
            except (ValueError, TypeError):
                # Valor no numérico, se ignora
                continue

    return wb, records, hes_column


def create_api_lookup(api_data: List[Dict[str, Any]]) -> Dict[int, Dict[str, Any]]:
    """
    Crea un diccionario de lookup por HESCode desde la respuesta de la API.

    Args:
        api_data: Lista de documentos de la API

    Returns:
        Dict {hes_code: documento}
    """
    from entities.boleta_response import extract_hes_code

    lookup = {}
    for doc in api_data:
        hes_code = extract_hes_code(doc)
        if hes_code is not None:
            lookup[hes_code] = doc

    return lookup


def process_records(excel_records: List[Dict[str, Any]], api_lookup: Dict[int, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Procesa los registros del Excel haciendo match con los datos de la API.

    Args:
        excel_records: Lista de registros del Excel con row_index y hes_code
        api_lookup: Diccionario {hes_code: documento de API}

    Returns:
        Lista de resultados con row_index, hes_code, boleta, detalle_errores, status
    """
    from entities.boleta_response import (
        extract_bte_code,
        extract_error_message,
        is_success,
    )

    results = []

    for record in excel_records:
        hes_code = record["hes_code"]
        row_index = record["row_index"]

        api_doc = api_lookup.get(hes_code)

        if api_doc is None:
            # No se encontró en la API
            error_raw = "NO_ENCONTRADO_EN_API"
            results.append({
                "row_index": row_index,
                "hes_code": hes_code,
                "boleta": 0,
                "detalle_errores": translate_error_message(error_raw),
                "status": "NOT_FOUND",
            })
        elif is_success(api_doc):
            # Exitoso
            bte_code = extract_bte_code(api_doc) or 0
            results.append({
                "row_index": row_index,
                "hes_code": hes_code,
                "boleta": bte_code,
                "detalle_errores": "",
                "status": "SUCCESS",
            })
        else:
            # Error
            error_raw = extract_error_message(api_doc) or "ERROR_DESCONOCIDO"
            results.append({
                "row_index": row_index,
                "hes_code": hes_code,
                "boleta": 0,
                "detalle_errores": translate_error_message(error_raw),
                "status": "ERROR",
            })

    return results


def find_column_by_name(worksheet, column_name: str) -> Optional[int]:
    """
    Busca una columna por nombre exacto en el header.

    Args:
        worksheet: Worksheet de openpyxl
        column_name: Nombre de la columna a buscar

    Returns:
        Índice de la columna (1-based) o None si no se encuentra
    """
    for cell in worksheet[1]:  # Primera fila (headers)
        if cell.value and str(cell.value).strip().upper() == column_name.upper():
            return cell.column
    return None


def write_output_excel(wb: Workbook, results: List[Dict[str, Any]], output_path: str):
    """
    Escribe el Excel de salida con las columnas BOLETA y DETALLE_ERRORES.
    Si las columnas ya existen, las sobrescribe. Si no, las crea al final.

    Args:
        wb: Workbook original
        results: Lista de resultados procesados
        output_path: Ruta donde guardar el Excel de salida
    """
    ws = wb.active

    # Buscar si las columnas ya existen
    boleta_col = find_column_by_name(ws, "BOLETA")
    detalle_col = find_column_by_name(ws, "DETALLE_ERRORES")

    # Si no existen, crearlas al final
    if boleta_col is None:
        boleta_col = ws.max_column + 1
        ws.cell(row=1, column=boleta_col, value="BOLETA")

    if detalle_col is None:
        detalle_col = ws.max_column + 1
        ws.cell(row=1, column=detalle_col, value="DETALLE_ERRORES")

    # Escribir los resultados
    for result in results:
        row_idx = result["row_index"]
        
        # Escribir BOLETA como número entero (sin formato de miles)
        boleta_cell = ws.cell(row=row_idx, column=boleta_col, value=result["boleta"])
        boleta_cell.number_format = '0'  # Formato de número sin decimales ni separadores
        
        # Escribir DETALLE_ERRORES como texto
        ws.cell(row=row_idx, column=detalle_col, value=result["detalle_errores"])

    # Guardar
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_output_filename(input_filename: str) -> str:
    """
    Genera el nombre del archivo de salida basado en el original.

    Args:
        input_filename: Nombre del archivo de entrada (ej: "reporte.xlsx")

    Returns:
        Nombre del archivo de salida con timestamp (ej: "reporte_procesado_20260206_150000.xlsx")
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_parts = Path(input_filename).stem
    extension = Path(input_filename).suffix

    return f"{name_parts}_procesado_{timestamp}{extension}"
